import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime, timedelta, time

# --- Funções Auxiliares para Cálculo de Tempo ---

def parse_time_to_timedelta(time_val):
    """Converte um valor de tempo (string HH:MM, datetime.time ou datetime.datetime) para um objeto timedelta."""
    if pd.isna(time_val):
        return None
    
    if isinstance(time_val, str):
        try:
            # Tenta converter diretamente de string HH:MM
            hours, minutes = map(int, time_val.split(":"))
            return timedelta(hours=hours, minutes=minutes)
        except ValueError:
            # Se falhar, tenta como datetime completo e extrai o tempo
            try:
                dt_obj = pd.to_datetime(time_val)
                return timedelta(hours=dt_obj.hour, minutes=dt_obj.minute)
            except Exception:
                return None
    elif isinstance(time_val, (datetime, time)):
        # Se já for um objeto datetime ou time, extrai horas e minutos
        return timedelta(hours=time_val.hour, minutes=time_val.minute)
    else:
        return None

def format_timedelta_to_hhmm(td):
    """Formata um objeto timedelta para uma string HH:MM."""
    if td is None:
        return ""
    total_seconds = int(td.total_seconds())
    if total_seconds < 0: 
        return "00:00" # Garante que não haja tempo negativo
    hours, remainder = divmod(total_seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}"

def calculate_time_overlap(start1_td, end1_td, start2_td, end2_td):
    """Calcula a sobreposição de dois intervalos de tempo (timedelta)."""
    if any(t is None for t in [start1_td, end1_td, start2_td, end2_td]):
        return timedelta(0)

    # Garante que os intervalos são válidos (fim >= inicio)
    # Se o fim for antes do início, o intervalo é inválido para cálculo de duração positiva
    if end1_td < start1_td or end2_td < start2_td:
        return timedelta(0) 

    # Encontra o início do período de sobreposição
    overlap_start = max(start1_td, start2_td)
    # Encontra o fim do período de sobreposição
    overlap_end = min(end1_td, end2_td)

    # Se houver sobreposição, calcula a duração
    if overlap_end > overlap_start:
        return overlap_end - overlap_start
    else:
        return timedelta(0)

# --- 1. CONFIGURAÇÃO E ESTILOS ---

# Caminhos
modelo_path = "modelo_incidentes_formatados.xlsx"
pasta_dados = "dados_xlsx"
arquivo_saida = "Incidentes_Formatados_Final.xlsx"

# Estilos base para aplicar nas novas linhas
altura_linha = 75
fonte_padrao = Font(name="Calibri", size=11)
alinhamento_celula = Alignment(horizontal="center", vertical="center", wrap_text=True)
borda_padrao = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# --- CABEÇALHOS DO SEU MODELO (ORDEM EXATA) ---
cabecalhos_modelo = [
    "LOJA", "REGIONAL", "ABERTURA", "FECHAMENTO", "CAUSA", 
    "TIPO", "STATUS", "LINK OPERANDO", "DATA", "INICIO", 
    "FIM", "IMPACTO", "DISPONIBILIDADE", "SOLUÇÃO"
]

# --- 2. CARREGAMENTO DO MODELO ---

try:
    # Carrega o workbook modelo
    wb = load_workbook(modelo_path)
    ws = wb.active
except FileNotFoundError:
    print(f"ERRO: Arquivo modelo \'{modelo_path}\' não encontrado.")
    exit()

# --- 3. PROCESSAMENTO E INSERÇÃO DOS DADOS (Arquivo por Arquivo) ---

# Define a linha inicial para inserção dos dados (linha 2, após o cabeçalho)
linha_atual = 2
print(f"Iniciando inserção de dados na planilha a partir da linha {linha_atual}...")

# Processa cada arquivo de dados na pasta
for nome_arquivo in os.listdir(pasta_dados):
    if nome_arquivo.endswith(".xlsx"):
        caminho_arquivo = os.path.join(pasta_dados, nome_arquivo)
        print(f"Lendo arquivo: {nome_arquivo}")

        # Lê a planilha com pandas
        df = pd.read_excel(caminho_arquivo)

        # --- APLICAÇÃO DA REGRA: REMOVER COLUNA "TITULO" ---
        if "Titulo" in df.columns:
            df = df.drop(columns=["Titulo"])
            print(f"  - Coluna \"Titulo\" removida do arquivo {nome_arquivo}. Dados deslocados para a esquerda.")

        # --- REORDENAR COLUNAS PARA CORRESPONDER AO MODELO ---
        # Cria um novo DataFrame com as colunas na ordem do modelo.
        # Se uma coluna do modelo não existir no df, ela será preenchida com NaN (vazio).
        df_reordenado = pd.DataFrame(columns=cabecalhos_modelo)
        for col in cabecalhos_modelo:
            if col in df.columns:
                df_reordenado[col] = df[col]
            else:
                df_reordenado[col] = None # Garante que a coluna exista e esteja vazia
        
        # --- APLICAÇÃO DA REGRA DE NEGÓCIO: "MATRIZ TI - Cuiabá" ---
        if not df_reordenado.empty:
            df_reordenado.loc[df_reordenado["LOJA"] == "MATRIZ TI - Cuiabá", ["ABERTURA", "FECHAMENTO"]] = ["06:00", "22:00"]
            print(f"  - Regra \"MATRIZ TI - Cuiabá\" aplicada no arquivo {nome_arquivo}.")

        # --- CÁLCULO DE IMPACTO E DISPONIBILIDADE ---
        # Itera sobre as linhas do DataFrame reordenado para calcular Impacto e Disponibilidade
        for index, row in df_reordenado.iterrows():
            abertura_td = parse_time_to_timedelta(row["ABERTURA"])
            fechamento_td = parse_time_to_timedelta(row["FECHAMENTO"])
            inicio_incidente_td = parse_time_to_timedelta(row["INICIO"])
            fim_incidente_td = parse_time_to_timedelta(row["FIM"])

            # Validação de dados irregulares (Fim < Inicio)
            if inicio_incidente_td is not None and fim_incidente_td is not None and fim_incidente_td < inicio_incidente_td:
                print(f"  - Dados de incidente irregulares (Fim < Início) na linha {index} do arquivo {nome_arquivo}. Impacto e Disponibilidade serão vazios.")
                df_reordenado.at[index, "IMPACTO"] = ""
                df_reordenado.at[index, "DISPONIBILIDADE"] = ""
                continue # Pula para a próxima linha se os dados forem irregulares

            # Cálculo da duração do expediente
            expediente_duracao = timedelta(0)
            if abertura_td is not None and fechamento_td is not None:
                expediente_duracao = fechamento_td - abertura_td

            # Cálculo do impacto
            impacto_duracao = timedelta(0)
            if inicio_incidente_td is not None and fim_incidente_td is not None and abertura_td is not None and fechamento_td is not None:
                # Limita o incidente ao horário de expediente
                incidente_inicio_efetivo = max(inicio_incidente_td, abertura_td)
                incidente_fim_efetivo = min(fim_incidente_td, fechamento_td)
                
                if incidente_fim_efetivo > incidente_inicio_efetivo:
                    impacto_duracao = incidente_fim_efetivo - incidente_inicio_efetivo

            # Cálculo da disponibilidade
            disponibilidade_duracao = expediente_duracao - impacto_duracao

            df_reordenado.at[index, "IMPACTO"] = format_timedelta_to_hhmm(impacto_duracao)
            df_reordenado.at[index, "DISPONIBILIDADE"] = format_timedelta_to_hhmm(disponibilidade_duracao)

        print(f"  - Cálculos de Impacto e Disponibilidade concluídos para o arquivo {nome_arquivo}.")

        # --- FILTRAR LINHAS COMPLETAMENTE VAZIAS ANTES DE ESCREVER ---
        # Considera uma linha vazia se todas as colunas, exceto IMPACTO e DISPONIBILIDADE, forem vazias
        colunas_para_verificar_vazio = [col for col in cabecalhos_modelo if col not in ["IMPACTO", "DISPONIBILIDADE"]]
        
        # Cria uma máscara booleana para identificar linhas que não são completamente vazias
        # Uma linha é considerada não vazia se pelo menos uma das colunas essenciais tiver um valor não nulo/não vazio
        # Preenche NaN com string vazia para que .strip() funcione e verifica se todas as colunas essenciais são vazias
        mask_empty_essential_cols = df_reordenado[colunas_para_verificar_vazio].apply(lambda x: x.astype(str).str.strip() == '').all(axis=1)
        df_reordenado_filtrado = df_reordenado[~mask_empty_essential_cols].copy()

        # --- INSERÇÃO E FORMATAÇÃO (com Openpyxl) ---
        # Itera sobre as linhas do DataFrame REORDENADO E FILTRADO
        for linha_dados in dataframe_to_rows(df_reordenado_filtrado, index=False, header=False):
            # Adiciona a linha de dados na posição correta
            # Garante que apenas as colunas definidas em cabecalhos_modelo sejam escritas
            for col_idx, valor_celula in enumerate(linha_dados[:len(cabecalhos_modelo)], 1):
                ws.cell(row=linha_atual, column=col_idx, value=valor_celula)
            
            # Aplica os estilos de formatação a essa nova linha
            ws.row_dimensions[linha_atual].height = altura_linha
            # Aplica estilo apenas às colunas que contêm dados (até o tamanho de cabecalhos_modelo)
            for col_num in range(1, len(cabecalhos_modelo) + 1):
                celula = ws.cell(row=linha_atual, column=col_num)
                celula.font = fonte_padrao
                celula.alignment = alinhamento_celula
                celula.border = borda_padrao
                
            # Incrementa o contador para a próxima linha
            linha_atual += 1

# --- LIMPEZA FINAL: REMOVER LINHAS VAZIAS NO FINAL DA PLANILHA ---
# Encontra a última linha com dados
max_row = ws.max_row
# Itera de baixo para cima, removendo linhas se todas as células estiverem vazias
while max_row > 1 and all(ws.cell(row=max_row, column=col).value is None for col in range(1, ws.max_column + 1)):
    ws.delete_rows(max_row)
    max_row -= 1

# --- 4. SALVAR O RESULTADO ---

try:
    wb.save(arquivo_saida)
    print(f"\n✅ Planilha final salva com sucesso como \'{arquivo_saida}\'")
except Exception as e:
    print(f"\nERRO ao salvar o arquivo final: {e}")



